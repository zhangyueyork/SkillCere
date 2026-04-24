#!/usr/bin/env python3
"""SkillCere command line utility.

First version: scan installed skills, show index status, and provide
recommendation context for an Agent.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
INDEX_FILE = ROOT / "skill-index.json"
PLATFORMS_FILE = ROOT / "platforms.json"
LOCAL_PLATFORMS_FILE = ROOT / "platforms.local.json"
INSTALL_LOG_FILE = ROOT / "install-log.jsonl"
SKILL_FILE_NAME = "SKILL.md"
DEFAULT_LOCAL_PLATFORMS = {
    "schema_version": "1.0.0",
    "updated_at": "",
    "platform_skill_dirs": {
        "codex": [
            {
                "path": str(Path.home() / ".codex" / "skills"),
                "role": "target",
                "write_policy": "read_write",
            }
        ],
        "agents": [
            {
                "path": str(Path.home() / ".agents" / "skills"),
                "role": "target",
                "write_policy": "read_write",
            }
        ],
        "claude-code": [
            {
                "path": str(Path.home() / ".claude" / "skills"),
                "role": "target",
                "write_policy": "read_write",
            }
        ],
        "gemini": [
            {
                "path": str(Path.home() / ".gemini" / "skills"),
                "role": "target",
                "write_policy": "read_write",
            }
        ],
    },
    "excluded_dirs": [],
}
SYNC_FILES = [
    INDEX_FILE,
    INSTALL_LOG_FILE,
    ROOT / "version-cache.json",
    PLATFORMS_FILE,
]


def now_iso() -> str:
    return datetime.now().astimezone().replace(microsecond=0).isoformat()


def read_json(path: Path, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"Missing required file: {path}")
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def ensure_project_files() -> None:
    if not INDEX_FILE.exists():
        write_json(
            INDEX_FILE,
            {
                "schema_version": "1.0.0",
                "updated_at": now_iso(),
                "skills": {},
                "aliases": {},
                "tags": {},
            },
        )
    if not PLATFORMS_FILE.exists():
        write_json(
            PLATFORMS_FILE,
            {
                "schema_version": "1.0.0",
                "platforms": {},
            },
        )
    if not (ROOT / "version-cache.json").exists():
        write_json(ROOT / "version-cache.json", {"schema_version": "1.0.0", "items": {}})
    if not INSTALL_LOG_FILE.exists():
        INSTALL_LOG_FILE.write_text("", encoding="utf-8")


def ensure_local_config() -> bool:
    if LOCAL_PLATFORMS_FILE.exists():
        return False
    local = DEFAULT_LOCAL_PLATFORMS.copy()
    local["updated_at"] = now_iso()
    write_json(LOCAL_PLATFORMS_FILE, local)
    return True


def append_log(event: dict[str, Any]) -> None:
    event = {"created_at": now_iso(), **event}
    with INSTALL_LOG_FILE.open("a", encoding="utf-8") as file:
        file.write(json.dumps(event, ensure_ascii=False) + "\n")


def relative_to_root(path: Path) -> str:
    return str(path.relative_to(ROOT)).replace("\\", "/")


def run_git(args: list[str], check: bool = False) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(
        ["git", *args],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )
    if check and result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()
        raise RuntimeError(detail or f"git {' '.join(args)} failed")
    return result


def sync_skill_index(
    message: str | None = None,
    *,
    dry_run: bool = False,
    push: bool = True,
) -> int:
    if run_git(["rev-parse", "--is-inside-work-tree"]).returncode != 0:
        print("Sync skipped: SkillCere is not inside a Git repository.")
        return 0

    existing_files = [path for path in SYNC_FILES if path.exists()]
    if not existing_files:
        print("Sync skipped: no syncable SkillCere files exist.")
        return 0

    relative_files = [relative_to_root(path) for path in existing_files]
    status = run_git(["status", "--porcelain", "--", *relative_files], check=True)
    changed = status.stdout

    if dry_run:
        print("SkillCere sync dry run")
        if changed.strip():
            print(changed.rstrip())
        else:
            print("No skill index changes to sync.")
        return 0

    if not changed.strip():
        print("No skill index changes to sync.")
        return 0

    run_git(["add", "--", *relative_files], check=True)
    staged = run_git(["diff", "--cached", "--name-only", "--", *relative_files], check=True)
    staged_files = [line.strip() for line in staged.stdout.splitlines() if line.strip()]
    if not staged_files:
        print("No skill index changes staged for sync.")
        return 0

    commit_message = message or "Sync SkillCere skill index"
    try:
        run_git(["commit", "-m", commit_message], check=True)
    except RuntimeError as error:
        print(f"Sync commit failed: {error}", file=sys.stderr)
        return 1

    remotes = [line.strip() for line in run_git(["remote"]).stdout.splitlines() if line.strip()]
    if not push:
        print("Skill index committed locally. Push skipped by --no-push.")
        return 0
    if not remotes:
        print("Skill index committed locally. No Git remote configured, so GitHub upload was skipped.")
        return 0

    try:
        run_git(["push"], check=True)
    except RuntimeError as error:
        print(f"Sync push failed: {error}", file=sys.stderr)
        return 1

    print("Skill index synced to Git remote.")
    return 0


def normalize_id(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9._-]+", "-", value)
    value = value.strip("-._")
    return value or "unknown-skill"


def parse_front_matter(text: str) -> tuple[dict[str, str], str]:
    if not text.startswith("---"):
        return {}, text

    lines = text.splitlines()
    end_index = None
    for index in range(1, len(lines)):
        if lines[index].strip() == "---":
            end_index = index
            break

    if end_index is None:
        return {}, text

    metadata: dict[str, str] = {}
    for line in lines[1:end_index]:
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip().strip("\"'")
        if key:
            metadata[key] = value

    body = "\n".join(lines[end_index + 1 :])
    return metadata, body


def first_heading(text: str) -> str:
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            return stripped.lstrip("#").strip()
    return ""


def first_paragraph(text: str) -> str:
    for block in re.split(r"\n\s*\n", text.strip()):
        block = block.strip()
        if not block or block.startswith("#") or block.startswith("```"):
            continue
        return re.sub(r"\s+", " ", block)[:300]
    return ""


def text_keywords(text: str, limit: int = 12) -> list[str]:
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_-]{2,}|[\u4e00-\u9fff]{2,}", text)
    stop_words = {
        "the",
        "and",
        "for",
        "with",
        "this",
        "that",
        "when",
        "use",
        "using",
        "skill",
        "skills",
        "should",
        "must",
        "will",
        "you",
        "your",
    }
    seen: set[str] = set()
    result: list[str] = []
    for token in tokens:
        normalized = token.lower()
        if normalized in stop_words or normalized in seen:
            continue
        seen.add(normalized)
        result.append(token)
        if len(result) >= limit:
            break
    return result


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return "sha256:" + digest.hexdigest()


def sha256_text(value: str, prefix: str = "sha256:") -> str:
    return prefix + hashlib.sha256(value.encode("utf-8")).hexdigest()


def is_excluded(path: Path, excluded_dirs: list[Path]) -> bool:
    try:
        resolved = path.resolve()
    except OSError:
        resolved = path
    for excluded in excluded_dirs:
        try:
            resolved.relative_to(excluded)
            return True
        except ValueError:
            continue
    return False


def discover_skill_dirs(root: Path, excluded_dirs: list[Path]) -> list[Path]:
    if not root.exists() or not root.is_dir():
        return []

    skills: list[Path] = []
    for skill_file in root.rglob(SKILL_FILE_NAME):
        skill_dir = skill_file.parent
        if is_excluded(skill_dir, excluded_dirs):
            continue
        skills.append(skill_dir)
    return sorted(set(skills), key=lambda item: str(item).lower())


def read_skill(skill_dir: Path) -> dict[str, Any]:
    skill_file = skill_dir / SKILL_FILE_NAME
    text = skill_file.read_text(encoding="utf-8", errors="replace")
    metadata, body = parse_front_matter(text)

    name = metadata.get("name") or skill_dir.name
    skill_id = normalize_id(name)
    description = metadata.get("description") or first_paragraph(body)
    if description in {"|", ">"}:
        description = first_paragraph(body)
    display_name = first_heading(body) or name
    file_hash = sha256_file(skill_file)
    keywords = text_keywords(" ".join([name, description, body[:2000]]))

    return {
        "id": skill_id,
        "name": name,
        "display_name": display_name,
        "description": description,
        "keywords": keywords,
        "categories": [],
        "source": {"type": "unknown", "url": ""},
        "latest_version": "unknown",
        "version_strategy": "file_hash",
        "latest_ref": file_hash,
        "status": "active",
    }


def ensure_index_shape(index: dict[str, Any]) -> dict[str, Any]:
    index.setdefault("schema_version", "1.0.0")
    index.setdefault("updated_at", now_iso())
    index.setdefault("skills", {})
    index.setdefault("aliases", {})
    index.setdefault("tags", {})
    return index


def merge_skill(index: dict[str, Any], skill: dict[str, Any], platform: str) -> str:
    skills = index["skills"]
    skill_id = skill["id"]
    existing = skills.get(skill_id)
    event_type = "unchanged"

    if existing is None:
        existing = skill | {
            "supported_platforms": [],
            "installed_on": {},
            "last_checked": "",
            "notes": "",
        }
        skills[skill_id] = existing
        event_type = "added"
    else:
        # Keep user-curated fields when present; only fill empty metadata.
        for key in ("name", "display_name", "description"):
            if not existing.get(key) and skill.get(key):
                existing[key] = skill[key]
        for key in ("source", "latest_version", "version_strategy", "latest_ref", "status"):
            existing.setdefault(key, skill.get(key))
        existing_keywords = set(existing.get("keywords", []))
        merged_keywords = list(existing.get("keywords", []))
        for keyword in skill.get("keywords", []):
            if keyword not in existing_keywords:
                merged_keywords.append(keyword)
                existing_keywords.add(keyword)
        existing["keywords"] = merged_keywords[:20]
        existing.setdefault("categories", [])
        existing.setdefault("supported_platforms", [])
        existing.setdefault("installed_on", {})

    if platform not in existing["supported_platforms"]:
        existing["supported_platforms"].append(platform)
        existing["supported_platforms"].sort()
        if event_type == "unchanged":
            event_type = "new_platform"

    previous_platform = existing["installed_on"].get(platform, {})
    if previous_platform.get("ref") != skill.get("latest_ref") and event_type == "unchanged":
        event_type = "updated"

    existing["installed_on"][platform] = {
        "version": skill.get("latest_version", "unknown"),
        "ref": skill.get("latest_ref", ""),
        "last_seen": now_iso(),
        "status": "installed",
    }
    return event_type


def aggregate_duplicate_skills(skills: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    refs: dict[str, set[str]] = {}

    for skill in skills:
        skill_id = skill["id"]
        if skill_id not in grouped:
            grouped[skill_id] = skill
            refs[skill_id] = set()
        refs[skill_id].add(skill.get("latest_ref", ""))

    for skill_id, skill_refs in refs.items():
        clean_refs = sorted(ref for ref in skill_refs if ref)
        if len(clean_refs) > 1:
            grouped[skill_id]["latest_ref"] = sha256_text("|".join(clean_refs), "sha256-multi:")

    return [grouped[key] for key in sorted(grouped)]


def load_scan_config() -> tuple[dict[str, Any], dict[str, Any], list[Path]]:
    ensure_project_files()
    created_local = ensure_local_config()
    if created_local:
        print(f"Created local scan config: {LOCAL_PLATFORMS_FILE}")
    platforms = read_json(PLATFORMS_FILE, {"platforms": {}})
    local = read_json(LOCAL_PLATFORMS_FILE, {"platform_skill_dirs": {}, "excluded_dirs": []})
    excluded = [Path(item).expanduser() for item in local.get("excluded_dirs", [])]
    return platforms, local, excluded


def command_scan(args: argparse.Namespace) -> int:
    index = ensure_index_shape(read_json(INDEX_FILE))
    platforms, local, excluded_dirs = load_scan_config()
    known_platforms = set(platforms.get("platforms", {}).keys())

    added = 0
    updated = 0
    unchanged = 0
    scanned = 0
    current_by_platform: dict[str, set[str]] = {}
    scanned_platforms: set[str] = set()
    missing_roots: list[str] = []

    for platform, roots in local.get("platform_skill_dirs", {}).items():
        if known_platforms and platform not in known_platforms:
            print(f"Warning: {platform} is not defined in platforms.json", file=sys.stderr)
        current_by_platform.setdefault(platform, set())

        for root_config in roots:
            if root_config.get("role", "target") not in {"target", "system"}:
                continue
            root = Path(root_config["path"]).expanduser()
            if not root.exists():
                missing_roots.append(f"{platform}: {root}")
                continue
            scanned_platforms.add(platform)
            discovered = [read_skill(skill_dir) for skill_dir in discover_skill_dirs(root, excluded_dirs)]
            for skill in aggregate_duplicate_skills(discovered):
                current_by_platform[platform].add(skill["id"])
                event_type = merge_skill(index, skill, platform)
                scanned += 1
                if event_type == "added":
                    added += 1
                elif event_type in {"updated", "new_platform"}:
                    updated += 1
                else:
                    unchanged += 1
                if event_type != "unchanged":
                    append_log(
                        {
                            "event": event_type,
                            "skill_id": skill["id"],
                            "platform": platform,
                            "version": skill.get("latest_version", "unknown"),
                            "ref": skill.get("latest_ref", ""),
                        }
                    )

    stale_platform_refs = 0
    stale_skills = 0
    for skill_id, skill in index["skills"].items():
        installed_on = skill.get("installed_on") or {}
        for platform in list(installed_on.keys()):
            if platform not in scanned_platforms:
                continue
            if skill_id not in current_by_platform.get(platform, set()):
                del installed_on[platform]
                stale_platform_refs += 1
                append_log(
                    {
                        "event": "not_seen",
                        "skill_id": skill_id,
                        "platform": platform,
                    }
                )
        skill["installed_on"] = installed_on
        if installed_on:
            skill["status"] = "active"
        else:
            if skill.get("status") != "not_installed":
                stale_skills += 1
                append_log({"event": "not_installed", "skill_id": skill_id})
            skill["status"] = "not_installed"

    index["updated_at"] = now_iso()
    write_json(INDEX_FILE, index)

    print(f"Scanned skills: {scanned}")
    print(f"Added skills: {added}")
    print(f"Updated or new platform sightings: {updated}")
    print(f"Unchanged sightings: {unchanged}")
    print(f"Stale platform sightings marked: {stale_platform_refs}")
    print(f"Skills marked not_installed: {stale_skills}")
    print(f"Total indexed skills: {len(index['skills'])}")
    if missing_roots:
        print("Missing scan roots:")
        for item in missing_roots:
            print(f"  - {item}")
    if not args.no_sync:
        print()
        sync_result = sync_skill_index()
        if sync_result != 0:
            return sync_result
    return 0


def command_status(_: argparse.Namespace) -> int:
    ensure_project_files()
    index = ensure_index_shape(read_json(INDEX_FILE))
    skills = index["skills"]
    platform_counts: dict[str, int] = {}
    missing_source = 0
    unknown_version = 0
    not_installed = 0

    for skill in skills.values():
        source = skill.get("source") or {}
        if not source.get("url"):
            missing_source += 1
        if skill.get("latest_version") in {"", None, "unknown"}:
            unknown_version += 1
        if skill.get("status") == "not_installed" or not (skill.get("installed_on") or {}):
            not_installed += 1
        for platform in (skill.get("installed_on") or {}).keys():
            platform_counts[platform] = platform_counts.get(platform, 0) + 1

    print("SkillCere status")
    print(f"  Indexed skills: {len(skills)}")
    print(f"  Installed skills: {len(skills) - not_installed}")
    print(f"  Not installed skills: {not_installed}")
    print(f"  Missing source URL: {missing_source}")
    print(f"  Unknown version: {unknown_version}")
    print("  Installed by platform:")
    if platform_counts:
        for platform, count in sorted(platform_counts.items()):
            print(f"    {platform}: {count}")
    else:
        print("    none")
    return 0


def command_prune(args: argparse.Namespace) -> int:
    ensure_project_files()
    index = ensure_index_shape(read_json(INDEX_FILE))
    skills = index["skills"]
    removable = [
        skill_id
        for skill_id, skill in skills.items()
        if skill.get("status") == "not_installed" or not (skill.get("installed_on") or {})
    ]
    removable.sort()

    if not args.drop_not_installed:
        print(f"Not installed skills eligible for pruning: {len(removable)}")
        for skill_id in removable[:50]:
            print(f"  - {skill_id}")
        if len(removable) > 50:
            print(f"  ... and {len(removable) - 50} more")
        print("Run with --drop-not-installed to delete these records from skill-index.json.")
        return 0

    for skill_id in removable:
        del skills[skill_id]
        append_log({"event": "pruned", "skill_id": skill_id})

    index["updated_at"] = now_iso()
    write_json(INDEX_FILE, index)
    print(f"Pruned not installed skills: {len(removable)}")
    print(f"Remaining indexed skills: {len(skills)}")
    return 0


def display_description(skill: dict[str, Any]) -> str:
    description = str(skill.get("description") or "").strip()
    if description in {"", "|", ">"}:
        return "暂无描述"
    return description


def compact_text(value: str, limit: int = 240) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def compact_skill_catalog(index: dict[str, Any], max_skills: int) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    for skill_id, skill in sorted(index.get("skills", {}).items()):
        installed_on = sorted((skill.get("installed_on") or {}).keys())
        catalog.append(
            {
                "id": skill_id,
                "name": skill.get("name", skill_id),
                "description": compact_text(display_description(skill), 260),
                "keywords": skill.get("keywords", [])[:12],
                "categories": skill.get("categories", [])[:8],
                "installed_on": installed_on,
                "version": skill.get("latest_version", "unknown"),
                "source_url_available": bool((skill.get("source") or {}).get("url")),
            }
        )
        if len(catalog) >= max_skills:
            break
    return catalog


def build_agent_context(task: str, index: dict[str, Any], max_skills: int) -> str:
    catalog = compact_skill_catalog(index, max_skills)
    return (
        "【SkillCere 推荐上下文】\n\n"
        "你是当前 Agent。SkillCere Core 只负责提供 skill 清单上下文，最终 skill 推荐应由你基于任务需求完成。\n\n"
        "请遵守：\n"
        "1. 主要推荐 skill，不要把平台选择作为主任务。\n"
        "2. 平台信息只用于说明 skill 已安装在哪里，或是否需要安装。\n"
        "3. 不要保存、复述或扩展用户隐私信息。\n"
        "4. 只从给定 skill 目录中推荐；如果没有合适 skill，要明确说没有。\n"
        "5. 推荐数量以 1-5 个为宜，宁缺毋滥。\n"
        "6. 对 source_url_available=false 或 version=unknown 的 skill，要说明目前无法确认远程最新版本，只能按本地 hash/状态追踪。\n\n"
        "请按以下格式输出：\n"
        "【推荐使用的 Skill】\n"
        "- skill-id：推荐理由；适用环节；安装平台提示\n\n"
        "【版本检查结果】\n"
        "- skill-id：版本状态，是否需要人工确认\n\n"
        "【安装/更新建议】\n"
        "- 说明当前平台是否可能需要安装这些 skill；不要给出不存在的下载地址\n\n"
        "【给执行 Agent 的启动说明】\n"
        "一段可以直接复制给执行 Agent 的说明，包含“使用哪些 skill 完成什么需求、执行顺序、缺失 skill 时如何处理”。\n\n"
        f"用户需求：\n{task}\n\n"
        "Skill 目录 JSON：\n"
        f"{json.dumps(catalog, ensure_ascii=False, indent=2)}\n"
    )


def command_context(args: argparse.Namespace) -> int:
    query = " ".join(args.task).strip()
    if not query:
        print("Please provide a task description.", file=sys.stderr)
        return 2

    ensure_project_files()
    index = ensure_index_shape(read_json(INDEX_FILE))
    print(build_agent_context(query, index, args.max_skills))
    return 0


def command_recommend(args: argparse.Namespace) -> int:
    print("Note: `recommend` is an alias of `context`; final recommendation should be done by the Agent.")
    print()
    return command_context(args)


def command_sync(args: argparse.Namespace) -> int:
    return sync_skill_index(
        args.message,
        dry_run=args.dry_run,
        push=not args.no_push,
    )


def command_export_excel(args: argparse.Namespace) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as error:
        print("openpyxl is required for export-excel.", file=sys.stderr)
        raise SystemExit(2) from error

    ensure_project_files()
    index = ensure_index_shape(read_json(INDEX_FILE))
    out_path = Path(args.output) if args.output else ROOT / "exports" / "skillcere-skill-list.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    platform_counts: dict[str, int] = {}
    missing_source = 0
    unknown_version = 0

    for skill_id, skill in sorted(index.get("skills", {}).items(), key=lambda item: item[0].lower()):
        if args.current_only and (skill.get("status") == "not_installed" or not (skill.get("installed_on") or {})):
            continue
        installed_on = skill.get("installed_on") or {}
        platforms = sorted(installed_on.keys())
        for platform in platforms:
            platform_counts[platform] = platform_counts.get(platform, 0) + 1
        source = skill.get("source") or {}
        source_url = source.get("url") or ""
        version = skill.get("latest_version") or "unknown"
        if not source_url:
            missing_source += 1
        if version == "unknown":
            unknown_version += 1
        last_seen_values = [
            value.get("last_seen", "")
            for value in installed_on.values()
            if isinstance(value, dict)
        ]
        rows.append(
            {
                "skill_id": skill_id,
                "name": skill.get("name", ""),
                "display_name": skill.get("display_name", ""),
                "description": skill.get("description", ""),
                "keywords": ", ".join(skill.get("keywords", []) or []),
                "categories": ", ".join(skill.get("categories", []) or []),
                "installed_on": ", ".join(platforms),
                "platform_count": len(platforms),
                "latest_version": version,
                "version_strategy": skill.get("version_strategy", ""),
                "source_type": source.get("type", ""),
                "source_url_available": "yes" if source_url else "no",
                "source_url": source_url,
                "last_checked": skill.get("last_checked", ""),
                "last_seen": max(last_seen_values) if last_seen_values else "",
                "status": skill.get("status", ""),
                "notes": skill.get("notes", ""),
            }
        )

    headers = [
        "skill_id",
        "name",
        "display_name",
        "description",
        "keywords",
        "categories",
        "installed_on",
        "platform_count",
        "latest_version",
        "version_strategy",
        "source_type",
        "source_url_available",
        "source_url",
        "last_checked",
        "last_seen",
        "status",
        "notes",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "skills"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, header in enumerate(headers, 1):
        letter = get_column_letter(column_index)
        if header == "description":
            width = 70
        elif header in {"keywords", "installed_on", "source_url"}:
            width = 36
        elif header in {"skill_id", "display_name"}:
            width = 28
        else:
            width = 18
        sheet.column_dimensions[letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if rows:
        table_ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName="SkillList", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    summary = workbook.create_sheet("summary")
    summary.append(["metric", "value"])
    summary_rows = [
        ("generated_at", now_iso()),
        ("indexed_skills", len(rows)),
        ("missing_source_url", missing_source),
        ("unknown_version", unknown_version),
        ("current_only", str(args.current_only).lower()),
    ]
    for item in summary_rows:
        summary.append(list(item))
    summary.append([])
    summary.append(["platform", "skill_count"])
    for platform, count in sorted(platform_counts.items()):
        summary.append([platform, count])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 40

    workbook.save(out_path)
    print(out_path)
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="skillcere",
        description="SkillCere: cross-Agent skill coordination utility.",
    )
    subcommands = parser.add_subparsers(dest="command", required=True)

    scan = subcommands.add_parser("scan", help="Scan local skill directories into the central index.")
    scan.add_argument("--no-sync", action="store_true", help="Do not sync the central index after scanning.")
    scan.set_defaults(func=command_scan)

    status = subcommands.add_parser("status", help="Show central index status.")
    status.set_defaults(func=command_status)

    prune = subcommands.add_parser("prune", help="Prune not-installed skills from the central index.")
    prune.add_argument("--drop-not-installed", action="store_true", help="Delete not-installed skill records.")
    prune.set_defaults(func=command_prune)

    context = subcommands.add_parser("context", help="Generate skill recommendation context for an Agent.")
    context.add_argument("task", nargs="+", help="Task description.")
    context.add_argument("--max-skills", type=int, default=300, help="Maximum number of indexed skills to include.")
    context.set_defaults(func=command_context)

    recommend = subcommands.add_parser("recommend", help="Alias of context.")
    recommend.add_argument("task", nargs="+", help="Task description.")
    recommend.add_argument("--max-skills", type=int, default=300, help="Maximum number of indexed skills to include.")
    recommend.set_defaults(func=command_recommend)

    sync = subcommands.add_parser("sync", help="Commit and push central skill index files.")
    sync.add_argument("-m", "--message", help="Git commit message.")
    sync.add_argument("--dry-run", action="store_true", help="Show syncable changes without committing.")
    sync.add_argument("--no-push", action="store_true", help="Commit locally but do not push.")
    sync.set_defaults(func=command_sync)

    export_excel = subcommands.add_parser("export-excel", help="Export the central index to an Excel workbook.")
    export_excel.add_argument("--output", help="Output .xlsx path.")
    export_excel.add_argument("--current-only", action="store_true", help="Export only currently installed skills.")
    export_excel.set_defaults(func=command_export_excel)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
